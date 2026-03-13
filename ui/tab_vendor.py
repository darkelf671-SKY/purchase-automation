"""탭 5: 업체 관리 — 업체 마스터 등록/수정/삭제 + Excel 일괄 업로드"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import db.vendor_repo as repo
from ui.design_system import COLORS, SPACING, FONTS, configure_treeview_tags, insert_with_alternating
from ui.base_dialog import BaseDialog

_PAY_LABELS = {"card": "법인카드", "transfer": "무통장입금", "auto_transfer": "자동이체납부"}


class VendorTab(ttk.Frame):
    def __init__(self, parent, status_var: tk.StringVar):
        super().__init__(parent, padding=SPACING["lg"])
        self.status_var = status_var
        self._vendors = []
        self._build_ui()

    def _build_ui(self):
        # ── 검색바 ──────────────────────────────────────────────
        search_frame = ttk.Frame(self)
        search_frame.pack(fill="x", pady=(0, SPACING["md"]))
        ttk.Label(search_frame, text="검색:").pack(side="left")
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._apply_filter())
        ttk.Entry(search_frame, textvariable=self._search_var,
                  width=22).pack(side="left", padx=SPACING["sm"])
        ttk.Button(search_frame, text="새로고침",
                   command=self.refresh).pack(side="left", padx=(SPACING["md"], 0))

        # 목록
        list_frame = ttk.LabelFrame(self, text=" 등록된 업체 ", padding=SPACING["lg"])
        list_frame.pack(fill="both", expand=True, pady=(0, SPACING["md"]))

        cols = ("상호", "대표자", "사업자번호", "결제방법", "주소")
        self._tree = ttk.Treeview(list_frame, columns=cols, show="headings", height=12)
        for col, w in zip(cols, [160, 90, 120, 100, 220]):
            self._tree.heading(col, text=col)
            self._tree.column(col, width=w, anchor="center")

        configure_treeview_tags(self._tree)

        btn_row = ttk.Frame(list_frame)
        btn_row.pack(side="bottom", fill="x", pady=(SPACING["sm"], 0))
        ttk.Button(btn_row, text="업체 추가", style="Primary.TButton",
                   command=self._add).pack(side="left", padx=SPACING["sm"])
        ttk.Button(btn_row, text="수정",
                   command=self._edit).pack(side="left", padx=SPACING["sm"])
        ttk.Button(btn_row, text="삭제", style="Danger.TButton",
                   command=self._delete).pack(side="right", padx=SPACING["sm"])
        ttk.Button(btn_row, text="Excel 일괄 업로드",
                   command=self._excel_upload).pack(side="left", padx=SPACING["sm"])
        ttk.Button(btn_row, text="양식 다운로드",
                   command=self._download_template).pack(side="left", padx=SPACING["sm"])

        sb = ttk.Scrollbar(list_frame, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=sb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._tree.bind("<Double-1>", lambda _: self._edit())

    def refresh(self):
        self._vendors = repo.select_all()
        self._apply_filter()

    def _apply_filter(self):
        kw = self._search_var.get().strip().lower()
        self._tree.delete(*self._tree.get_children())
        for v in self._vendors:
            if kw and not any(kw in str(v.get(f, "")).lower()
                              for f in ("name", "ceo", "business_no", "address", "bank_name")):
                continue
            pay = _PAY_LABELS.get(v["payment_method"], v["payment_method"])
            insert_with_alternating(self._tree, "", "end", iid=str(v["id"]), values=(
                v["name"], v["ceo"], v["business_no"], pay, v["address"]
            ))

    def _get_selected(self) -> dict | None:
        sel = self._tree.selection()
        if not sel:
            messagebox.showwarning("선택 오류", "업체를 선택하세요.")
            return None
        vid = int(sel[0])
        return next((v for v in self._vendors if v["id"] == vid), None)

    def _add(self):
        VendorDialog(self, title="업체 추가", on_save=self._on_save)

    def _edit(self):
        v = self._get_selected()
        if v:
            VendorDialog(self, title="업체 수정", vendor=v, on_save=self._on_save)

    def _delete(self):
        v = self._get_selected()
        if not v:
            return
        if messagebox.askyesno("삭제 확인", f"'{v['name']}' 업체를 삭제하시겠습니까?"):
            try:
                repo.delete(v["id"])
            except Exception:
                messagebox.showerror("삭제 오류",
                    f"'{v['name']}' 업체가 구매 이력에서 참조 중이므로 삭제할 수 없습니다.")
                return
            self.refresh()
            self.status_var.set(f"'{v['name']}' 삭제 완료")

    def _on_save(self):
        self.refresh()
        self.status_var.set("업체 정보 저장 완료")

    # ── Excel 일괄 업로드 ─────────────────────────────────────

    def _download_template(self):
        """업체 일괄 등록용 Excel 양식 다운로드 (예제 + 설명 시트 포함)"""
        path = filedialog.asksaveasfilename(
            title="업체 일괄등록 양식 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile="업체_일괄등록_양식.xlsx",
        )
        if not path:
            return

        wb = Workbook()

        # ── 시트1: 입력 양식 ──
        ws = wb.active
        ws.title = "업체목록"

        headers = ["상호 *", "대표자", "사업자등록번호", "주소",
                   "은행명", "예금주", "계좌번호"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 예제 데이터 (2행)
        examples = [
            ["(주)한솔사무용품", "김철수", "123-45-67890", "서울시 강남구 테헤란로 123",
             "", "", ""],
            ["오피스디포", "이영희", "234-56-78901", "서울시 종로구 종로 456",
             "국민은행", "오피스디포", "123-456-789012"],
            ["에스투비", "박지성", "345-67-89012", "서울시 서초구 서초대로 789",
             "", "", ""],
        ]
        example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        for r, row_data in enumerate(examples, 2):
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.fill = example_fill
                cell.border = thin_border

        # 열 너비 조정
        widths = [22, 12, 18, 35, 12, 12, 20]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

        # ── 시트2: 설명 ──
        ws2 = wb.create_sheet("설명")
        ws2.sheet_properties.tabColor = "FFC000"

        info = [
            ["업체 일괄 등록 양식 설명", ""],
            ["", ""],
            ["필드명", "설명"],
            ["상호 *", "필수 입력. 업체명 (예: (주)한솔사무용품)"],
            ["대표자", "선택. 대표자 이름"],
            ["사업자등록번호", "선택. 중복 검사 기준 (예: 123-45-67890)"],
            ["주소", "선택. 업체 주소"],
            ["은행명", "선택. 무통장입금 업체 시 입력 (예: 국민은행)"],
            ["예금주", "선택. 무통장입금 업체 시 입력 (예: 홍길동)"],
            ["계좌번호", "선택. 무통장입금 업체 시 입력 (예: 123-456-789012)"],
            ["", ""],
            ["결제방법 자동 감지 규칙", ""],
            ["", "• 은행명 + 계좌번호 모두 입력 → 무통장입금으로 자동 인식"],
            ["", "• 은행명 또는 계좌번호만 입력 → 법인카드로 인식 (불완전 정보 경고)"],
            ["", "• 둘 다 비어있으면 → 법인카드로 자동 설정"],
            ["", "• 자동이체납부 업체는 프로그램에서 직접 등록하세요"],
            ["", ""],
            ["중복 처리 규칙", ""],
            ["", "사업자등록번호가 같은 업체가 이미 등록되어 있으면 '중복'으로 표시됩니다."],
            ["", "사업자등록번호가 없으면 상호명으로 중복 검사합니다."],
            ["", "중복 업체는 '건너뛰기' 또는 '덮어쓰기(업데이트)' 중 선택 가능합니다."],
            ["", ""],
            ["주의사항", ""],
            ["", "• '업체목록' 시트의 1행(헤더)은 수정하지 마세요."],
            ["", "• 예제 데이터(초록 행)는 삭제하고 실제 데이터를 입력하세요."],
        ]
        title_font = Font(bold=True, size=14, color="1F4E79")
        section_font = Font(bold=True, size=11, color="2E75B6")
        for r, (a, b) in enumerate(info, 1):
            ws2.cell(row=r, column=1, value=a)
            ws2.cell(row=r, column=2, value=b)
            if r == 1:
                ws2.cell(row=r, column=1).font = title_font
            elif a and not b:
                ws2.cell(row=r, column=1).font = section_font
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 65

        wb.save(path)
        self.status_var.set(f"양식 저장 완료: {Path(path).name}")
        messagebox.showinfo("양식 다운로드", f"양식이 저장되었습니다.\n{path}")

    def _excel_upload(self):
        """Excel 파일로 업체 일괄 등록"""
        path = filedialog.askopenfilename(
            title="업체 일괄등록 Excel 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xls")]
        )
        if not path:
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb.active

            header_map = {
                "상호": "name", "상호 *": "name",
                "대표자": "ceo",
                "사업자등록번호": "business_no",
                "주소": "address",
                "결제방법": "_legacy_payment",  # 이전 양식 하위 호환
                "은행명": "bank_name",
                "예금주": "account_holder",
                "계좌번호": "account_no",
            }
            # 헤더 매핑
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                val = str(cell.value or "").strip()
                headers.append(header_map.get(val))

            if "name" not in headers:
                messagebox.showerror("양식 오류",
                    "첫 행에 '상호' 또는 '상호 *' 헤더가 필요합니다.\n"
                    "'양식 다운로드' 버튼으로 올바른 양식을 받으세요.")
                return

            # 데이터 파싱
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                data = {
                    "name": "", "ceo": "", "business_no": "", "address": "",
                    "bank_name": "", "account_holder": "", "account_no": "",
                    "is_auto_transfer": 0,
                }
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        data[headers[i]] = str(val or "").strip()
                if not data["name"]:
                    continue
                # 이전 양식 하위 호환: _legacy_payment 제거
                data.pop("_legacy_payment", None)
                # payment_method는 vendor_repo.insert/bulk_insert에서 자동 계산
                rows.append(data)
            wb.close()

            if not rows:
                messagebox.showinfo("데이터 없음", "등록할 업체 데이터가 없습니다.")
                return

        except Exception as e:
            messagebox.showerror("파일 오류", f"Excel 파일을 읽을 수 없습니다.\n{e}")
            return

        # 중복 검사 + 미리보기
        BulkUploadPreviewDialog(self, rows, on_save=self._on_bulk_save)

    def _on_bulk_save(self):
        self.refresh()
        # 구매 조사 탭 업체 목록도 갱신
        try:
            notebook = self.master
            for tab_id in notebook.tabs():
                tab_widget = notebook.nametowidget(tab_id)
                if hasattr(tab_widget, 'refresh_vendors'):
                    tab_widget.refresh_vendors()
                    break
        except Exception:
            pass


class VendorDialog(BaseDialog):
    def __init__(self, parent, title: str, vendor: dict = None, on_save=None):
        self._vendor = vendor
        self._vars = {}
        self._auto_transfer_var = tk.IntVar(value=0)
        self._pay_preview_var = tk.StringVar(value="→ 기본 결제방법: 법인카드")
        super().__init__(parent, title, on_save=on_save)
        if vendor:
            self._load(vendor)

    def _build_content(self, f: ttk.Frame):
        fields = [
            ("상호 *:", "name", 30),
            ("대표자:", "ceo", 20),
            ("사업자등록번호:", "business_no", 20),
            ("주소:", "address", 40),
        ]
        for r, (lbl, key, w) in enumerate(fields):
            ttk.Label(f, text=lbl).grid(row=r, column=0, sticky="w",
                                        pady=SPACING["sm"], padx=(0, SPACING["md"]))
            var = tk.StringVar()
            ttk.Entry(f, textvariable=var, width=w).grid(row=r, column=1, columnspan=3,
                                                         sticky="ew", pady=SPACING["sm"])
            self._vars[key] = var

        # 구분선
        ttk.Separator(f, orient="horizontal").grid(
            row=4, column=0, columnspan=4, sticky="ew", pady=SPACING["md"])

        # 은행 정보 (항상 표시)
        bank_fields = [
            ("은행명:", "bank_name", 16),
            ("예금주:", "account_holder", 16),
            ("계좌번호:", "account_no", 24),
        ]
        for r, (lbl, key, w) in enumerate(bank_fields, start=5):
            ttk.Label(f, text=lbl).grid(row=r, column=0, sticky="w", pady=SPACING["sm"])
            var = tk.StringVar()
            var.trace_add("write", lambda *_: self._update_pay_preview())
            ttk.Entry(f, textvariable=var, width=w).grid(
                row=r, column=1, columnspan=3, sticky="w", pady=SPACING["sm"])
            self._vars[key] = var

        # 자동이체 체크박스
        ttk.Checkbutton(f, text="자동이체납부 업체",
                        variable=self._auto_transfer_var,
                        command=self._update_pay_preview).grid(
            row=8, column=0, columnspan=4, sticky="w", pady=SPACING["sm"])

        # 자동 감지 미리보기
        ttk.Label(f, textvariable=self._pay_preview_var,
                  foreground=COLORS["primary"],
                  font=("맑은 고딕", 9, "bold")).grid(
            row=9, column=0, columnspan=4, sticky="w", pady=SPACING["sm"])

        f.columnconfigure(1, weight=1)

    def _update_pay_preview(self):
        """은행 정보/자동이체에 따라 기본 결제방법 미리보기 갱신"""
        from db.vendor_repo import derive_payment_method
        bank = self._vars.get("bank_name", tk.StringVar()).get()
        acct = self._vars.get("account_no", tk.StringVar()).get()
        is_at = bool(self._auto_transfer_var.get())
        result = derive_payment_method(bank, acct, is_at)
        label = _PAY_LABELS.get(result, result)
        self._pay_preview_var.set(f"→ 기본 결제방법: {label}")

    def _load(self, v: dict):
        for key, var in self._vars.items():
            var.set(v.get(key, ""))
        self._auto_transfer_var.set(v.get("is_auto_transfer", 0) or 0)
        self._update_pay_preview()

    def _on_save(self):
        name = self._vars["name"].get().strip()
        if not name:
            messagebox.showwarning("입력 오류", "상호를 입력하세요.", parent=self)
            return

        bank_name = self._vars["bank_name"].get().strip()
        account_no = self._vars["account_no"].get().strip()

        # 불완전 은행 정보 경고
        from db.vendor_repo import validate_bank_info
        warning = validate_bank_info(bank_name, account_no)
        if warning:
            messagebox.showwarning("은행 정보 확인", warning, parent=self)

        data = {
            "name":             name,
            "ceo":              self._vars["ceo"].get().strip(),
            "business_no":      self._vars["business_no"].get().strip(),
            "address":          self._vars["address"].get().strip(),
            "bank_name":        bank_name,
            "account_holder":   self._vars["account_holder"].get().strip(),
            "account_no":       account_no,
            "is_auto_transfer": self._auto_transfer_var.get(),
            # payment_method는 vendor_repo에서 자동 계산
        }
        try:
            if self._vendor:
                repo.update(self._vendor["id"], data)
            else:
                repo.insert(data)
            self._fire_save_callback()
            self.destroy()
        except Exception as e:
            msg = str(e)
            if "UNIQUE" in msg.upper():
                msg = f"'{name}' 업체명이 이미 등록되어 있습니다."
            messagebox.showerror("저장 오류", msg, parent=self)


class BulkUploadPreviewDialog(tk.Toplevel):
    """일괄 업로드 미리보기 — 중복 업체 처리 선택"""

    _PAY = {"card": "법인카드", "transfer": "무통장입금", "auto_transfer": "자동이체납부"}

    def __init__(self, parent, rows: list[dict], on_save=None):
        super().__init__(parent)
        self.title("업체 일괄 등록 미리보기")
        self.geometry("850x520")
        self.transient(parent)
        self.grab_set()
        self._rows = rows
        self._on_save = on_save
        self._dup_actions: dict[int, tk.StringVar] = {}
        self._build_ui()

    def _build_ui(self):
        # 상단 요약
        summary_f = ttk.Frame(self, padding=SPACING["md"])
        summary_f.pack(fill="x")

        # 중복 검사
        self._duplicates: list[tuple[int, dict, dict]] = []  # (idx, new_row, existing)
        new_count = 0
        for i, row in enumerate(self._rows):
            existing = None
            biz_no = row.get("business_no", "").strip()
            if biz_no:
                existing = repo.find_by_business_no(biz_no)
            if not existing:
                existing = repo.find_by_name(row["name"])
            if existing:
                self._duplicates.append((i, row, existing))
            else:
                new_count += 1
                row["_action"] = "insert_new"

        dup_count = len(self._duplicates)
        ttk.Label(summary_f, text=f"총 {len(self._rows)}건",
                  font=FONTS["heading"]).pack(side="left", padx=(0, SPACING["lg"]))
        ttk.Label(summary_f, text=f"신규: {new_count}건",
                  foreground=COLORS["success"]).pack(side="left", padx=SPACING["sm"])
        ttk.Label(summary_f, text=f"중복: {dup_count}건",
                  foreground=COLORS["danger"] if dup_count else COLORS["text_secondary"]
                  ).pack(side="left", padx=SPACING["sm"])

        if dup_count:
            # 일괄 처리 버튼
            batch_f = ttk.Frame(summary_f)
            batch_f.pack(side="right")
            ttk.Button(batch_f, text="모두 건너뛰기",
                       command=lambda: self._set_all_action("skip")).pack(side="left", padx=SPACING["xs"])
            ttk.Button(batch_f, text="모두 덮어쓰기",
                       command=lambda: self._set_all_action("update")).pack(side="left", padx=SPACING["xs"])

        # 테이블
        tree_f = ttk.Frame(self, padding=(SPACING["md"], 0))
        tree_f.pack(fill="both", expand=True)

        cols = ("상태", "상호", "대표자", "사업자번호", "결제방법", "처리")
        self._tree = ttk.Treeview(tree_f, columns=cols, show="headings", height=16)
        for col, w in zip(cols, [70, 160, 90, 130, 100, 100]):
            self._tree.heading(col, text=col)
            self._tree.column(col, width=w, anchor="center")

        configure_treeview_tags(self._tree)
        self._tree.tag_configure("dup", background="#FFF3CD")
        self._tree.tag_configure("new", background="#D4EDDA")

        sb = ttk.Scrollbar(tree_f, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=sb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # 데이터 채우기
        for i, row in enumerate(self._rows):
            is_dup = any(d[0] == i for d in self._duplicates)
            from db.vendor_repo import derive_payment_method
            auto_pay = derive_payment_method(
                row.get("bank_name", ""),
                row.get("account_no", ""),
                bool(row.get("is_auto_transfer", 0)),
            )
            pay = self._PAY.get(auto_pay, auto_pay)
            tag = "dup" if is_dup else "new"
            status = "중복" if is_dup else "신규"
            action = "건너뛰기" if is_dup else "등록"
            self._tree.insert("", "end", iid=str(i), values=(
                status, row["name"], row["ceo"], row["business_no"], pay, action
            ), tags=(tag,))

        # 중복 건 처리 선택
        if dup_count:
            dup_f = ttk.LabelFrame(self, text=" 중복 업체 처리 ", padding=SPACING["md"])
            dup_f.pack(fill="x", padx=SPACING["md"], pady=(SPACING["sm"], 0))

            ttk.Label(dup_f,
                      text="사업자등록번호 또는 상호가 이미 등록된 업체입니다. 처리 방법을 선택하세요.",
                      foreground=COLORS["text_secondary"]).pack(anchor="w")

            dup_scroll = ttk.Frame(dup_f)
            dup_scroll.pack(fill="x", pady=(SPACING["sm"], 0))

            for idx, new_row, existing in self._duplicates:
                row_f = ttk.Frame(dup_scroll)
                row_f.pack(fill="x", pady=SPACING["xs"])
                match_field = "사업자번호" if (new_row.get("business_no") and
                    existing.get("business_no") == new_row["business_no"]) else "상호"
                ttk.Label(row_f,
                          text=f"  '{new_row['name']}' ({match_field} 일치: '{existing['name']}')",
                          width=50, anchor="w").pack(side="left")
                var = tk.StringVar(value="skip")
                self._dup_actions[idx] = var
                new_row["_action"] = "skip"
                ttk.Radiobutton(row_f, text="건너뛰기", variable=var, value="skip",
                                command=lambda i=idx: self._update_action(i)).pack(side="left")
                ttk.Radiobutton(row_f, text="덮어쓰기(업데이트)", variable=var, value="update",
                                command=lambda i=idx: self._update_action(i)).pack(side="left")

        # 하단 버튼
        btn_f = ttk.Frame(self, padding=SPACING["md"])
        btn_f.pack(fill="x")
        ttk.Button(btn_f, text="취소", command=self.destroy).pack(side="right", padx=SPACING["sm"])
        ttk.Button(btn_f, text="일괄 등록 실행", style="Primary.TButton",
                   command=self._execute).pack(side="right", padx=SPACING["sm"])

    def _update_action(self, idx: int):
        action = self._dup_actions[idx].get()
        self._rows[idx]["_action"] = action
        label = "덮어쓰기" if action == "update" else "건너뛰기"
        self._tree.set(str(idx), "처리", label)

    def _set_all_action(self, action: str):
        for idx, var in self._dup_actions.items():
            var.set(action)
            self._rows[idx]["_action"] = action
            label = "덮어쓰기" if action == "update" else "건너뛰기"
            self._tree.set(str(idx), "처리", label)

    def _execute(self):
        result = repo.bulk_insert(self._rows)
        msg = (
            f"일괄 등록 완료\n\n"
            f"  신규 등록: {result['inserted']}건\n"
            f"  업데이트: {result['updated']}건\n"
            f"  건너뜀: {result['skipped']}건"
        )
        messagebox.showinfo("일괄 등록 결과", msg, parent=self)

        if self._on_save:
            self._on_save()
        self.destroy()
